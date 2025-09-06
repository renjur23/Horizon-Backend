from rest_framework import viewsets, filters
from rest_framework.permissions import IsAuthenticated,IsAuthenticatedOrReadOnly
from authentication.permissions import IsAdminOrEmployeeCanCreate
from .models import (
    Client, Location, InverterStatus, Inverter, Generator,
    SiteContact, Order, InverterSimDetail, InverterUtilizationStatus,
    InverterUtilization, ServiceStatus, ServiceRecords, Usage, 
    Checklist, ChecklistItem, BatteryVoltage
)
from .serializers import *
from authentication.permissions import IsAdminUser, AdminOnlyFieldsPermission
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from .models import Usage, Inverter, Order
import pandas as pd
from decimal import Decimal
from datetime import datetime
from rest_framework import status
import os
from rest_framework.parsers import MultiPartParser
from django.core.mail import send_mail
from django.conf import settings
import openpyxl
from django.utils.timezone import make_aware, is_naive
from django.db.models import Count
from django.shortcuts import get_object_or_404
from django.db.models import Sum,FloatField,F
from rest_framework.decorators import action
from django.utils.timezone import now
import logging
from django.db.models import Q
from rest_framework.pagination import PageNumberPagination


logger = logging.getLogger(__name__)

class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

class LocationViewSet(viewsets.ModelViewSet):
    queryset = Location.objects.all()
    serializer_class = LocationSerializer
    permission_classes = [IsAuthenticated]

class InverterStatusViewSet(viewsets.ModelViewSet):
    queryset = InverterStatus.objects.all()
    serializer_class = InverterStatusSerializer
    permission_classes = [IsAuthenticated]
    
    def create(self, request, *args, **kwargs):
        print("AUTH HEADER:", request.headers.get('Authorization'))  # ✅ Check if token is received
        return super().create(request, *args, **kwargs)



class InverterViewSet(viewsets.ModelViewSet):
    queryset = Inverter.objects.all()
    serializer_class = InverterSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['unit_id', 'model', 'serial_no', 'inverter_status__inverter_status_name','given_name']
    
    
    def get_queryset(self):
        queryset = super().get_queryset()
        status_param = self.request.query_params.get("status")
        if status_param:
            status_list = [s.strip() for s in status_param.split(',') if s.strip()]
            q = Q()
            for s in status_list:
                q |= Q(inverter_status__inverter_status_name__iexact=s)
            queryset = queryset.filter(q)
        return queryset


class GeneratorViewSet(viewsets.ModelViewSet):
    queryset = Generator.objects.all()
    serializer_class = GeneratorSerializer
    permission_classes = [IsAuthenticated]



class SiteContactViewSet(viewsets.ModelViewSet):
    queryset = SiteContact.objects.all()
    serializer_class = SiteContactSerializer
    permission_classes = [IsAuthenticated]


class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all().order_by('po_number')
    lookup_field = 'id' 
    permission_classes = [IsAuthenticated, IsAdminOrEmployeeCanCreate, AdminOnlyFieldsPermission]

    def get_serializer_class(self):
        if self.action == 'create':
            return OrderCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return OrderUpdateSerializer
        return OrderSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()

        # ✅ Filter by status if it's in query params
        status = self.request.query_params.get("status")
        if status:
            queryset = queryset.filter(status__iexact=status)

        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        # 🔑 Return full OrderSerializer so React sees location_name, generator_no, site_contact_name
        read_serializer = OrderSerializer(instance, context=self.get_serializer_context())
        return Response(read_serializer.data)
    
    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        print(data)

        # ✅ Handle Location
        if data.get("location"):
            location_obj, _ = Location.objects.get_or_create(
                location_name=data.get("location")
            )
            data["location_id"] = location_obj.id

        # ✅ Handle Generator
        if all(
            [
                data.get("generator_no"),
                data.get("fuel_consumption"),
                data.get("generator_size"),
            ]
        ):
            generator_obj, _ = Generator.objects.get_or_create(
                generator_no=data.get("generator_no"),
                generator_size=data.get("generator_size"),
                fuel_consumption=data.get("fuel_consumption"),
            )
            data["generator_no"] = generator_obj.id
        if all(
            [
                data.get("site_contact_name"),
                data.get("site_contact_email"),
                data.get("site_contact_number"),
            ]
        ):
            site_contact_obj, _ = SiteContact.objects.get_or_create(
                site_contact_name=data.get("site_contact_name"),
                site_contact_email=data.get("site_contact_email"),
                site_contact_number=data.get("site_contact_number"),
            )
            data["site_contact_id"] = site_contact_obj.id


        # ✅ Create serializer with modified data
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        inverter=Inverter.objects.get(id=data["inverter_id"])
        inverter.inverter_status=InverterStatus.objects.get(inverter_status_name="Hired")
        inverter.save()
        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )
        
    @action(detail=True, methods=['post'], url_path='offhire')
    def offhire(self, request, id=None):
        order = self.get_object()

        order.end_date = now().date()
        order.save()

        if order.inverter_id:
            testing_status = InverterStatus.objects.get(inverter_status_name="Testing")
            order.inverter_id.inverter_status = testing_status
            order.inverter_id.save()

        inverter_name = "N/A"
        if order.inverter_id:
            inverter_parts = [
                order.inverter_id.given_start_name or "",
                order.inverter_id.model or "",
                order.inverter_id.serial_no or "",
            ]
            inverter_name = " ".join(filter(None, inverter_parts))
            
            offhired_by_name = getattr(request.user, "name", request.user.username)
            offhired_by_email = getattr(request.user, "email", "N/A")

        # ✅ Email content
        subject = 'Inverter Unit Offhired'
        message = f"""
    The following inverter unit has been offhired:

    Unit          : {inverter_name}
    PO Number     : {order.po_number}
    Contract No   : {order.contract_no}
    Client        : {order.issued_to.client_name if order.issued_to else 'N/A'}
    End Date      : {order.end_date}
    Location      : {getattr(order.location_id, 'location_name', 'N/A')}
    Remarks       : {order.remarks or 'None'}
    Offhired By   : {offhired_by_name} ({offhired_by_email})
    """

        from_email = settings.DEFAULT_FROM_EMAIL
        recipient_list = [
            # 'steevo@offgridenergy.ie',
            # 'Jp@generatorhire.ie',
            # 'anna@offgridenergy.ie',
            # 'David@horizonplant.com',
            # 'swathy.horizonoffgrid@gmail.com',
            # 'john@generatorhire.ie',
            'renjurenjithrajendran@gmail.com'
        ]

        if order.issued_to and order.issued_to.client_email:
            recipient_list.append(order.issued_to.client_email)

        try:
            send_mail(subject, message, from_email, recipient_list, fail_silently=False)
            logger.info(f"✅ Offhire email sent for PO {order.po_number}, unit {inverter_name}")
        except Exception as e:
            logger.error(f"❌ Error sending offhire email for PO {order.po_number}: {str(e)}")

        return Response({"message": f"Order offhired successfully. Email sent for unit {inverter_name}."},
                        status=status.HTTP_200_OK)




class InverterSimDetailViewSet(viewsets.ModelViewSet):
    queryset = InverterSimDetail.objects.all().order_by('inverter_id__unit_id')
    serializer_class = InverterSimDetailSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['phone_number', 'serial_no', 'inverter_id__given_name', 
    'inverter_id__serial_no', ]
    
   


class InverterUtilizationStatusViewSet(viewsets.ModelViewSet):
    queryset = InverterUtilizationStatus.objects.all()
    serializer_class = InverterUtilizationStatusSerializer
    permission_classes = [IsAuthenticated]


class InverterUtilizationViewSet(viewsets.ModelViewSet):
    queryset = InverterUtilization.objects.all()
    serializer_class = InverterUtilizationSerializer
    permission_classes = [IsAuthenticated]


class ServiceStatusViewSet(viewsets.ModelViewSet):
    queryset = ServiceStatus.objects.all()
    serializer_class = ServiceStatusSerializer
    permission_classes = [IsAuthenticated]


class ServiceRecordsViewSet(viewsets.ModelViewSet):
    queryset = ServiceRecords.objects.all()
    serializer_class = ServiceRecordsSerializer
    permission_classes = [IsAuthenticated]


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 100
    page_size_query_param = "page_size"
    max_page_size = 500

class UsageViewSet(viewsets.ModelViewSet):
    serializer_class = UsageSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        qs = Usage.objects.all().order_by('-date')
        po_number = self.request.query_params.get("po_number")
        inverter_id = self.request.query_params.get("inverter_id")
        from_date = self.request.query_params.get("from_date")
        to_date = self.request.query_params.get("to_date")

        if po_number:
            qs = qs.filter(order_id__po_number__iexact=po_number)
        if inverter_id:
            qs = qs.filter(inverter_id=inverter_id)
        if from_date:
            qs = qs.filter(date__gte=from_date)
        if to_date:
            qs = qs.filter(date__lte=to_date)
        return qs



class UsageUploadView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response({"error": "No file provided"}, status=400)

        # --- Read Excel into DataFrame ---
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet_name = "For Access" if "For Access" in wb.sheetnames else wb.sheetnames[0]
            df = pd.read_excel(excel_file, engine="openpyxl", sheet_name=sheet_name)
            df.columns = df.columns.str.strip()
        except Exception as e:
            return Response({"error": f"File read error: {str(e)}"}, status=400)

        # --- Column mapping ---
        column_mapping = {
            "Inverter_unit_id": "inverter_unit_id",
            "po_number": "po_number",
            "date": "date",
            "kw_consumed": "kw_consumed",
            "generator_run_hour": "generator_run_hour",
        }
        df.rename(columns=column_mapping, inplace=True)

        required_columns = {"inverter_unit_id", "po_number", "date", "kw_consumed", "generator_run_hour"}
        if not required_columns.issubset(df.columns):
            return Response({"error": f"Missing required columns. Required: {required_columns}"}, status=400)

        # --- Prefetch lookups ---
        inverter_map = {inv.unit_id: inv for inv in Inverter.objects.all()}
        order_map = {o.po_number.lower(): o for o in Order.objects.all()}

        # --- Prefetch existing usages (avoid duplicates) ---
        unique_keys = set()
        for _, row in df.iterrows():
            unit_id = str(row["inverter_unit_id"]).strip()
            po_number = str(row["po_number"]).strip().lower() if not pd.isna(row["po_number"]) else None
            date_value = pd.to_datetime(row["date"], errors="coerce")
            if unit_id and date_value:
                unique_keys.add((unit_id, po_number, date_value.date()))

        existing_usages = set(
            Usage.objects.filter(
                inverter_id__unit_id__in=[k[0] for k in unique_keys],
                order_id__po_number__in=[k[1] for k in unique_keys if k[1]]
            ).values_list("inverter_id__unit_id", "order_id__po_number", "date")
        )

        # --- Process rows ---
        skipped_rows, new_usages = [], []
        success_count = 0

        for index, row in df.iterrows():
            try:
                unit_id = str(row["inverter_unit_id"]).strip()
                po_number = str(row["po_number"]).strip().lower() if not pd.isna(row["po_number"]) else None
                inverter = inverter_map.get(unit_id)

                if not inverter:
                    skipped_rows.append(f"⚠️ Inverter not found: {unit_id}")
                    continue

                try:
                    date_value = pd.to_datetime(row["date"], errors="coerce")
                    if pd.isna(date_value):
                        raise ValueError("Invalid date")
                    if is_naive(date_value):
                        date_value = make_aware(date_value)

                    kw_consumed = float(row["kw_consumed"])
                    gen_hr = float(row["generator_run_hour"])
                    site_hr = float(row.get("site_run_hour", 24))
                except Exception:
                    skipped_rows.append(f"❌ Invalid data format at row {index + 2}: {row.to_dict()}")
                    continue

                order = order_map.get(po_number) if po_number else None

                if (unit_id, po_number, date_value.date()) in existing_usages:
                    continue  # Skip duplicates

                usage = Usage(
                    inverter_id=inverter,
                    order_id=order,
                    is_yard=False,
                    date=date_value,
                    kw_consumed=kw_consumed,
                    generator_run_hour=gen_hr,
                    site_run_hour=site_hr,
                    inverter_usage_calculated=str((24 - gen_hr) / 24),
                    generator_run_hour_save=str(site_hr - gen_hr),
                    inverter_usage_based_on_site_run_hour=str((site_hr - gen_hr) / site_hr),
                    inverter_usage_based_on_site=str((site_hr - gen_hr) / site_hr),
                )
                new_usages.append(usage)
                success_count += 1

            except Exception as e:
                skipped_rows.append(f"❌ Error in row {index + 2}: {str(e)}")

        # --- Bulk insert new records ---
        if new_usages:
            Usage.objects.bulk_create(new_usages, ignore_conflicts=True)

        return Response({
            "message": f"{success_count} records processed successfully.",
            "skipped_rows": skipped_rows,
        }, status=201)

class InverterUsageReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, inverter_id):
        inverter = get_object_or_404(Inverter, id=inverter_id)
        orders = Order.objects.filter(inverter_id=inverter)

        if not orders.exists():
            return Response({"error": "No order found for this inverter"}, status=404)

        order = orders.first()
        usage_qs = Usage.objects.filter(inverter_id=inverter, order_id=order).order_by("date")

        fuel_price = order.fuel_price or 1.25
        co2_per_litre = order.co2_emission_per_litre or 2.678
        fuel_consumption = order.generator_no.fuel_consumption if order.generator_no else 12

        daily_usage = []
        total_kw = 0
        total_hr_saved = 0
        total_fuel_saved = 0
        total_cost_saved = 0
        total_co2 = 0

        for u in usage_qs:
            saved_hr = u.site_run_hour - u.generator_run_hour
            battery_pct = round((saved_hr / u.site_run_hour) * 100, 2) if u.site_run_hour else 0.0
            fuel_saved = round(saved_hr * fuel_consumption, 2)
            savings = round(fuel_saved * fuel_price, 2)
            co2 = round(fuel_saved * co2_per_litre, 2)

            daily_usage.append({
                "date": u.date,
                "kw_consumed": u.kw_consumed,
                "generator_run_hour": u.generator_run_hour,
                "generator_run_hour_save": round(saved_hr, 2),
                "inverter_usage_based_on_site": battery_pct,
                "fuel_saved": fuel_saved,
                "fuel_cost_saved": savings,
                "co2_saved": co2,
            })

            total_kw += u.kw_consumed
            total_hr_saved += saved_hr
            total_fuel_saved += fuel_saved
            total_cost_saved += savings
            total_co2 += co2

        avg_usage_percent = round(
            sum([d["inverter_usage_based_on_site"] for d in daily_usage]) / len(daily_usage), 2
        ) if daily_usage else 0.0

        return Response({
            "po_number": order.po_number,
            "client_name": order.issued_to.client_name if order.issued_to else "",
            "location_name": order.location_id.location_name if order.location_id else "",
            "generator_no": {
                "generator_no": order.generator_no.generator_no if order.generator_no else "",
                "fuel_consumption": fuel_consumption
            },
            "inverter_unit_id": inverter.unit_id,
            "inverter_given_name": inverter.given_name,
            "fuel_price": fuel_price,
            "co2_per_litre": co2_per_litre,
            "total_kw": total_kw,
            "total_hr_saved": round(total_hr_saved, 2),
            "total_fuel_saved": round(total_fuel_saved, 2),
            "fuel_cost_saved": round(total_cost_saved, 2),
            "co2_saved": round(total_co2, 2),
            "avg_usage_percent": avg_usage_percent,
            "daily_usage": daily_usage,
        })
        
class ChecklistViewSet(viewsets.ModelViewSet):
    queryset = Checklist.objects.all().order_by('created_at')
    serializer_class = ChecklistSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]

    def create(self, request, *args, **kwargs):
        """
        Custom create:
        - Accept nested items + batteries
        - Return detailed response with nested serializers
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        checklist = serializer.save()  # ✅ uses nested create logic in ChecklistSerializer
        read_serializer = self.get_serializer(checklist)

        return Response(read_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        """
        Custom update:
        - Handle updating checklist, items, and batteries
        """
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        checklist = serializer.save()
        read_serializer = self.get_serializer(checklist)

        return Response(read_serializer.data, status=status.HTTP_200_OK)
    
    
class InverterStatusSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        status_counts = (
            Inverter.objects
            .values('inverter_status__inverter_status_name')
            .annotate(count=Count('id'))
        )

        summary = {
            entry['inverter_status__inverter_status_name']: entry['count']
            for entry in status_counts
        }
        return Response(summary)
    
    
